def main():
    import pandas as pd
    from getkey import getkey,key
    import numpy as np
    import re 
    import openpyxl
    from openpyxl.styles import PatternFill,Alignment,NamedStyle
    from openpyxl.styles.borders import Border, Side
    from openpyxl.styles import Font, Color
    from openpyxl.styles import Border,Side
    import ascii_magic
    import openpyxl
    import operator
    from functools import reduce

    try:
        from openpyxl.cell import get_column_letter
    except ImportError:
        from openpyxl.utils import get_column_letter
        from openpyxl.utils import column_index_from_string

    from openpyxl import load_workbook
    from openpyxl import Workbook
    import os
    import shutil 
    import time

    try:
        # current kaam chalu hai


        # to add the logo of c5i
        output = ascii_magic.from_image_file("C5-Logo.jpg",column=1000,char='#')
        ascii_magic.to_terminal(output)

        # to check the excel file format
        def open_file(file_path):
            if '.xlsx' in file_path:

                # to open the excel file to get l0 

        #         df = pd.read_excel(file_path)
        #         # print(df)
        #         l0 = list(df.columns)
        #         # print(l0) 
        #         global len_of_lst1
        #         len_of_lst1 = len(l0) + 3 
        #         # print (l0)

                # to open the excel file to get l1 automatically



                for iterate_sheet in range(len(ws_total_sheets)):

                    wb = openpyxl.load_workbook(file_path,data_only=True)
                    ws = wb[ws_total_sheets[iterate_sheet]] 

                    df = pd.read_excel(file_path,sheet_name=ws_total_sheets[iterate_sheet])
                    # print(df)
                    l0 = list(df.columns)
                    # print(l0) 
                    global len_of_lst1
                    len_of_lst1 = len(l0) + 3 
    #                 print (l0)

                    l1 = []
                    first_row = ws[2] 
                    for cell in first_row:
                        l1.append('%s' % (cell.value))  
    #                 print(l1) 

                    # REgex
                    regex_regex_float = '[+-]?[0-9]+\.[0-9]+'
                    regex_for_special_string = '[@_!#$%.^&*()<>?/\|}{~:]'
                    regex_for_percentage = '\\d+(\\.?\\d+)?%'
                    regex_for_numbers_commas_minus = '^([-+] ?)?[0-9]+(,[0-9]+)?$'
                    regex_for_alphanum = "^[A-Za-z0-9_ -]*$"   # aadil space given for including space in alpha numeric

                    # to get the data type of the l1
                    dt = dict() 
                    a_true = ""
                    a_true_two = ""  #
                    j_needed_value = ""
                    j_needed_value_two = "" #
                    # print(len(l0), 'len of the l0')
                    # print(len(l1), 'len of the l1')

                    # to check the date
                    for j in range(len(l0)):
                        if 'month' in l0[j].lower():

    #                         print(l1[j])
                            j_needed_value = j_needed_value + str(j) 
                            a_true = a_true + 'True'
                            a = l1[j]
                            zzz = j
                            l = len(a)
                            if ("'") in a and l == 6:
                                try:
                                    dt[zzz] = "Month-(date)"
    #                                 print('bakkal')
                                except:
                                    print('  ') 

                            elif ('00:00:00' in a) or (len(a) == 8) or (len(a) == 10) or ('%-%-%' or '%/%/%' in a):
                                dt[zzz] = "Date-(Month)"



                        elif 'date' in l0[j].lower():
    #                         print('haan bhai date toh le rha h yaar yeh') 
    #                         print(j,l0[j].lower())
                            j_needed_value_two = j_needed_value_two + str(j)  #
                            a_true_two = a_true_two + 'True' #
                            a = l1[j]
                            zzz = j
                            # print(j)
    #                         print(len(a)) 
                            if '00:00:00' in a:
                                a.replace(' 00:00:00','')
    #                         print(len(a)) 
                            if len(a) == 8:
    #                             print('8')
                                dt[zzz] = "Date"
                            elif len(a) == 10:
    #                             print('10')
                                dt[zzz] = "Date"
                            elif '%-%-%' or '%/%/%' in a:
                                dt[zzz] = "Date"
    #                             print('date with dash or /')
                            elif '00:00:00' in a:
    #                             print('hao bhai')
                                dt[zzz] = "Date" 





                    # logic:- if the column is not a date 
                    if len(j_needed_value)> 0:
                        j_needed_value1 = int(j_needed_value)
                    else:
                        j_needed_value1 = j_needed_value

                    if len(j_needed_value_two)> 0: #
                        j_needed_value_two1 = int(j_needed_value_two)
                    else: #
                        j_needed_value_two1 = j_needed_value_two
                    # print(j_needed_value1)
                    # print(int(float(j_needed_value)))
                    # print(len(j_needed_value)) 
                    # print(type(j_needed_value1))  
                    # print(a_true)

                    global semi_col 
                    for i in range(len(l1)):
                        # print(l1[i])
                        if i == j_needed_value1 or i == j_needed_value_two1:
                            semi_col= l1[i].count(":")                         # to check if the count of the decimal in the string is 2, if so then it's a time.
                            # print(semi_col,'semi')
                            if a_true == 'True':
                                i = i + 1 

                                if l1[i].isnumeric() == True:
                                    dt[i] = "Integer"                     
                                elif (re.search(regex_for_percentage,l1[i])):
                                    dt[i] = "Percentage"
                                elif (re.search(regex_for_numbers_commas_minus,l1[i])):
                                    dt[i] = "Integer"


                                elif (l1[i].find('®') or l1[i].find('™')) != -1:
                                    dt[i] = "String"
                                elif l1[i].isalpha() == True or l1[i].isalnum() == True:
                                    dt[i] = "String"
                                elif l1[i] == '/' or l1[i] == 'None' or l1[i] == 'N/A':
                                    dt[i] = '--'
                                elif (re.search(regex_for_special_string,l1[i])):
                                    if semi_col == 2 and len(l1[i])==8:
                                        dt[i] = "Time" 
                                    else:
                                        dt[i] = "String"
                                elif (re.search(regex_regex_float,l1[i])):
                                    dt[i] = "Float"
                                elif (re.search(regex_for_alphanum,l1[i])):
                                    dt[i] = 'String' 

                                else:
                                    dt[i] = "--"

                            '''---------------------------------'''
                        semi_col= l1[i].count(":")
                        if l1[i].isnumeric() == True:
                            dt[i] = "Integer"                     
                        elif (re.search(regex_for_percentage,l1[i])):
                            dt[i] = "Percentage"
                        elif (re.search(regex_for_numbers_commas_minus,l1[i])):
                            dt[i] = "Integer"


                        elif (l1[i].find('®') or l1[i].find('™')) != -1:
                            dt[i] = "String"
                        elif l1[i].isalpha() == True or l1[i].isalnum() == True:
                            dt[i] = "String"
                        elif l1[i] == '/' or l1[i] == 'None' or l1[i] == 'N/A':
                            dt[i] = '--'
                        elif (re.search(regex_for_special_string,l1[i])):
                            if semi_col == 2 and len(l1[i])==8:
                                dt[i] = "Time"
                            else:
                                dt[i] = "String"
                        elif (re.search(regex_regex_float,l1[i])):
                            dt[i] = "Float"
                        elif (re.search(regex_for_alphanum,l1[i])):
                            dt[i] = 'String'

                        else:
                            dt[i] = "--"

                    non_sorted_dict = dict(dt.items())
                    # print(non_sorted_dict)

                    sorted_by_key_dict = {k:v for k,v in sorted(dt.items())}
                    # print(sorted_by_key_dict)

            #         ct = dict(sorted(dt.items()))
                    # print(dt.sort())
                    a = list(sorted_by_key_dict.items())
                    lst_of_sorted_items = []  


                    out = [item for t in a for item in t]
                    for i in range(0,len(out)):
                        # print(out[i])
                        if i%2!=0 :
                            lst_of_sorted_items.append(out[i])
                    # print(lst_of_sorted_items)
                    numpy_lst1 = np.array(lst_of_sorted_items)

                    df1=pd.DataFrame(columns=['Table',"KPI's","Datatype","Sample Data"])
                    df1["KPI's"] = pd.Series(l0)
                    df1["Sample Data"] = pd.Series(l1)
                    df1["Datatype"] = pd.Series(numpy_lst1)


                    # print(df1)

                    data_schema_file_name_extractor = file_path.split('\\').pop()
                    final_file_value = data_schema_file_name_extractor.split('.')[0]
                    global final_name1
                    final_name1 = f"Data Schema -{final_file_value}[{ws_total_sheets[iterate_sheet]}]- Tech Community.xlsx"
                    df1.loc[0,'Table'] = final_file_value

                    df1.to_excel(final_name1,index=False,index_label=False)

                    writer = pd.ExcelWriter(final_name1, engine='xlsxwriter')
                    header = pd.MultiIndex.from_product([[f"Data Schema -{final_file_value}- Tech Community"],df1.columns])
                    df1 = pd.DataFrame(df1.to_numpy(), index=df1.index , columns = header) 
                    # df.to_excel(writer,sheet_name=f'Sheet1',index=True,index_label=True)
                    final_new_name = f"Data Schema - {final_file_value}[{ws_total_sheets[iterate_sheet]}] - Tech Community.xlsx"

                    df1.reset_index(drop=True, inplace=True)

                    df1.to_excel(final_new_name)



                    wb = openpyxl.load_workbook(final_new_name)
                    ws = wb['Sheet1'] 

                    fill_pattern = PatternFill(patternType='solid', fgColor='BDD7EE')

                    ws['B2'].fill = fill_pattern
                    ws['C2'].fill = fill_pattern
                    ws['D2'].fill = fill_pattern
                    ws['E2'].fill = fill_pattern
                    ws["B4"].alignment = Alignment(horizontal="center",vertical='center')
                    ws.merge_cells(f"B4:B{len_of_lst1}") 


                    wb.save(final_new_name) 



                    wb = openpyxl.load_workbook(final_new_name)
                    ws = wb['Sheet1']
                    # print('1')

                    top=Side(border_style='thin',color="000000")
                    bottom=Side(border_style='thin', color="000000")
                    right_side = Side(border_style='thin', color="000000")
                    left_side = Side(border_style='thin', color="000000")
                    border=Border(top=top,bottom=bottom,right=right_side,left=left_side)

                    font_style1 = Font(size = 14,bold=True)
                    font_style2 = Font(size = 13,bold=True)
                    font_style3 = Font(size=12)
                    font_style4 = Font(size=12,bold=True)


                    # print('2')
                    for i in range(4,len_of_lst1+1):
                        ws.cell(row=i, column=5).font = font_style3
                        ws.cell(row=i,column=4).font = font_style3
                        ws.cell(row=i, column=3).font = font_style3
                        ws.cell(row=i,column=2).font = font_style1 #
                        ws.cell(row=i, column=5).border = border
                        ws.cell(row=i,column=4).border = border
                        ws.cell(row=i, column=3).border = border
                        ws.cell(row=i,column=2).border = border

                    for i in range(1,6):
                        ws.cell(row=2,column=i).font = font_style2
                        ws.cell(row=1,column=i).font = font_style1
                        ws.cell(row=2,column=i).border = border
                        ws.cell(row=1,column=i).border = border

                    # Imorting the necessary modules
                    # stackoverflow code to automatic change the width of the columns




                    for column_cells in ws.columns:
                        new_column_length = max(len(str(cell.value)) for cell in column_cells)
                        new_column_letter = (get_column_letter(column_cells[0].column))
                        if new_column_length > 0:
                            ws.column_dimensions[new_column_letter].width = new_column_length*1.32

                    # for i in range(4,len_of_lst+1):
                    #     ws.cell(row=i, column=5).border = border
                    #     ws.cell(row=i,column=4).border = border
                    #     ws.cell(row=i, column=3).border = border
                    #     ws.cell(row=i,column=2).border = border
                    # for i in range(1,6):
                    #     ws.cell(row=2,column=i).border = border
                    #     ws.cell(row=1,column=i).border = border
    #                 ws.column_dimensions['A'].hidden = True
    #                 ws.row_dimensions[3].hidden = True

    #         to change the date's data-type:- 
                    values_of_kpi = [c.value for c in ws['C'][3:]]
                    values_of_data_types = [d.value for d in ws['D'][3:]]
                    values_of_first_column = [e.value for e in ws['E'][3:]] 

                    len_of_first_column = len(values_of_first_column)

                    # print(values_of_data_types)
                    # print(len(values_of_data_types))
                    # print('==================')
                    # print(values_of_first_column)
                    # print(len(values_of_first_column))
                    # print('==================')
                    # print(values_of_kpi)
                    # print(len(values_of_kpi))
                    # print('===================')
    #                 1,len(a)+1
                    for finding_dates in range(0,len_of_first_column):
                        if 'date' in values_of_kpi[finding_dates].lower():
                            a = finding_dates

    #                           a = 0
                    #                         print(a)  
                            if '00:00:00' in values_of_first_column[finding_dates]:

                                values_of_first_column[finding_dates] = values_of_first_column[finding_dates].replace('00:00:00','') 
                                val1 = values_of_first_column[a]
                                ws.cell(row= 4+a,column=5).value = val1
                            if len(values_of_first_column[finding_dates])==8:
                                values_of_data_types[finding_dates] = 'Date'
                                val = values_of_data_types[a]
                                ws.cell(row= 4+a,column=4).value = val 

                            elif len(values_of_first_column[finding_dates])==10:
                                values_of_data_types[finding_dates] = 'Date'
                                val = values_of_data_types[a]
                                ws.cell(row= 4+a,column=4).value = val
                            elif '%-%-%' or '%/%/%' in values_of_first_column[finding_dates]:
                                values_of_data_types[finding_dates] = 'Date'
                                val = values_of_data_types[a]
                                ws.cell(row= 4+a,column=4).value = val 

                        if values_of_first_column[finding_dates].lower() == 'none' or values_of_first_column[finding_dates].lower() == 'nan':
                            a2 = finding_dates

                            values_of_data_types[a2] = '--' 
                            val = values_of_data_types[a2]
                            ws.cell(row= 4+a2,column=4).value = val

                        # if 'month' in values_of_kpi[finding_dates].lower() and '00:00:00' in values_of_first_column[finding_dates]:
                        #     a3 = finding_dates
                        #     values_of_data_types[a3] = 'Date-[Month]'
                        #     val2 = values_of_data_types[a3] 
                        #     ws.cell(row=4+a3,column=4).value = val2 
                        



                    # print(values_of_data_types)
                    # print(len(values_of_data_types))
                    # print('==================')
                    # print(values_of_first_column)
                    # print(len(values_of_first_column))
                    # print('==================')
                    # print(values_of_kpi)
                    # print(len(values_of_kpi))
                    # print('===================') 



                    # ws.title = f"DS-{ws_total_sheets[iterate_sheet]}" 
                    ws.title = f"Data-Schema" 
                    ws.column_dimensions['A'].hidden = True
                    ws.row_dimensions[3].hidden = True
                    wb.save(final_new_name)

                    # import os
                    # with open(filename, 'x') as f:
                    # os.remove(final_name)
    #                 print(ws_total_sheets[iterate_sheet])                           <---------------------------------

                    global secondry_sheet_intermediate_file_remover
                    secondry_sheet_intermediate_file_remover = f"Data Schema -{final_file_value}[{ws_total_sheets[0]}]- Tech Community.xlsx"

                    # to move our file to D-disk:-
                    pwd_for_source = os.getcwd()
                    source_file_path = os.path.join(pwd_for_source,final_new_name)
                    destination_file_path = os.path.join("C:\\C5i-Data Schema Files",final_new_name)
                    shutil.move(source_file_path, destination_file_path) 



        # create directory in D drive
        dir = os.path.join("C:\\","C5i-Data Schema Files")
        if not os.path.exists(dir):
            os.mkdir(dir) 

        file_path = input('enter your file path: ')
        if file_path[0] == '"' and file_path[-1] == '"':
            file_path = file_path[1:]
            file_path = file_path[:-1]

        wb = openpyxl.load_workbook(file_path,data_only=True)      # <----------------------------------
        ws_total_sheets = wb.sheetnames
    #     print(ws_total_sheets,'jjjjjjjjjjjjjjjjjjjjjjjj')


        open_file(file_path) 



        pwd1 = os.getcwd()
        # print(pwd1)
        file_path = pwd1+ f'\{final_name1}'
        file_path1 = pwd1+ f'\{final_name1}' 


        if os.path.isfile(file_path):
            os.remove(file_path)
            print("\n \nDone!")
        elif os.path.isfile(file_path):
            os.remove(file_path)
            print("\n\nDone!") 
        else:
            print("Intermediate File does not exist - DON'T WORRY IT's NOT An Error")    

        remove_extra_file_name_when_sheets_are_2 = pwd1 + f"\{secondry_sheet_intermediate_file_remover}"

        if os.path.isfile(remove_extra_file_name_when_sheets_are_2):
            os.remove(remove_extra_file_name_when_sheets_are_2)
            print("\n")
        elif os.path.isfile(remove_extra_file_name_when_sheets_are_2):
            os.remove(remove_extra_file_name_when_sheets_are_2)
            print("\n") 
        else:
            print("Intermediate File does not exist - Don't Worry It's Not an Error")

            print('Thanks for Using Schema Creator')
            print("\n\n                                   Press Enter To Create More Schemas...")
            # time.sleep(1)
            
            
            var = getkey()

            if var == key.ENTER:
                main()

            else:
#                 time.sleep(10)
                exit()



    except Exception as e:
        print('The Exact Error is Available below:-')
        print(e)
        print('\n') 
        print('Pls Contact Team:Course5i to report the Bug')
        time.sleep(10)

    else:
    #     print('Thanks for Using Schema Creator')  
        time.sleep(1)
        
main() 
